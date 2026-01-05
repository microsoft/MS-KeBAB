import json
import re
from collections import defaultdict
from collections.abc import Callable
from pathlib import Path
from typing import cast

import numpy as np
import pytest
from kebab import mskebab
from openpyxl import load_workbook
from openpyxl.cell.cell import Cell


def assert_dicts_almost_equal(dict1, dict2, tol=1e-6):
    """Assert that two dictionaries with numeric values are almost equal."""
    assert dict1.keys() == dict2.keys(), "Dictionaries have different keys"
    for key in dict1:
        val1 = dict1[key]
        val2 = dict2[key]
        if isinstance(val1, dict) and isinstance(val2, dict):
            assert_dicts_almost_equal(val1, val2, tol)
        else:
            if val1 is None and val2 is None:
                continue
            if (val1 is None and val2 is not None) or (val1 is not None and val2 is None):
                pytest.fail(f"Values for key '{key}' differ: {val1} vs {val2}")
            assert abs(val1 - val2) < tol, f"Values for key '{key}' differ: {val1} vs {val2}"


def tokenize(formula: str) -> list[str]:
    """Tokenize a formula string into components."""
    tokens = []
    current_token = ""
    for char in formula:
        if char in "+-*/(),":
            if current_token:
                tokens.append(current_token)
                current_token = ""
            tokens.append(char)
        elif char.isspace():
            if current_token:
                tokens.append(current_token)
                current_token = ""
        else:
            current_token += char
    if current_token:
        tokens.append(current_token)
    return tokens


def is_float(token: str) -> bool:
    """Check if a token can be converted to a float."""
    try:
        float(token)
        return True
    except ValueError:
        return False


def is_cell_or_range_reference(token: str) -> bool:
    """Check if a token is a cell or range reference (e.g., A1, B2:D4)."""
    # Simple check: starts with letters followed by numbers, possibly with a colon for ranges
    if ":" in token:
        parts = token.split(":")
        return all(part[0].isalpha() and part[1:].isdigit() for part in parts)
    return token[0].isalpha() and token[1:].isdigit()


def split_arguments(arg_string: str) -> list[str]:
    """Split a string of arguments by commas, ignoring commas within quoted strings and parentheses."""
    args = []
    current_arg = ""
    in_quotes = False
    in_parentheses = False
    for char in arg_string:
        if char == '"':
            in_quotes = not in_quotes
            current_arg += char
        elif char == "," and not in_quotes and not in_parentheses:
            args.append(current_arg.strip())
            current_arg = ""
        elif char == "{" and not in_quotes:
            in_parentheses = True
            current_arg += char
        elif char == "}" and not in_quotes:
            in_parentheses = False
            current_arg += char
        else:
            current_arg += char
    if current_arg:
        args.append(current_arg.strip())
    return args


class MetricsSpreadsheet:
    def __init__(self, debug_output_file: str) -> None:
        debug_workbook = load_workbook(debug_output_file, data_only=False)
        self.debug_info_sheet = debug_workbook["debug_info"]
        self.evaluated_cells = {}
        self.evaluators = {
            "SUMIFS": self.evaluate_sumifs,
            "AVERAGEIF": self.evaluate_averageif,
        }

    def convert_cell_value(self, cell: Cell) -> float | str | None:
        if cell.value is None:
            return None
        if is_float(str(cell.value)):
            return float(cell.value)  # type: ignore
        return str(cell.value)

    def get_range_values(self, range_part: str) -> list[float | str | None]:
        range_start, range_end = range_part.split(":")
        cells = cast(tuple[tuple[Cell]], self[range_start:range_end])
        return [self.convert_cell_value(cell) for row in cells for cell in row]

    def parse_criteria(self, criteria: str) -> list[str]:
        """Parse criteria if needed."""
        if "{" in criteria and "}" in criteria:
            # Handle array criteria if needed
            criteria = criteria.strip()[1:-1]
            return [c.strip().strip('"') for c in split_arguments(criteria)]
        return [criteria.strip().strip('"')]

    def evaluate_criterion(self, value: float | str, criterion: str) -> bool:
        property_id_pattern = r"[a-zA-Z0-9_ ,]+"
        if criterion == "?*":
            return len(str(value)) > 0
        if criterion.startswith("<>"):
            return str(value) != criterion[2:]
        if re.match(property_id_pattern, criterion):
            return str(value) == criterion
        if criterion.startswith((">", "<", ">=", "<=")) and is_float(str(value)):
            return eval(f"{value}{criterion}" if not np.isnan(value) else f"np.{value}{criterion}")
        return False

    def evaluate_criteria(self, value: float | str, criteria: str) -> bool:
        parsed_criteria = self.parse_criteria(criteria)
        return all(self.evaluate_criterion(value, crit) for crit in parsed_criteria)

    def evaluate_sumifs(self, sumif_content: str) -> float:
        """Evaluate a SUMIFS function given its content and a spreadsheet context."""
        tokens = split_arguments(sumif_content)
        sum_range_part = tokens[0].strip()
        sum_range = self.get_range_values(sum_range_part)
        sum_range_floats = [float(value) if value is not None else 0.0 for value in sum_range]
        criteria_ranges = tokens[1::2]
        criteria = tokens[2::2]

        criteria_indices = set(range(len(sum_range)))
        for range_part, criterion in zip(criteria_ranges, criteria, strict=True):
            range_ = self.get_range_values(range_part.strip())
            criterion_ = criterion.strip().strip('"')
            criterion_indices = set()
            for idx, value in enumerate(range_):
                if value is not None and self.evaluate_criteria(value, criterion_):
                    criterion_indices.add(idx)
            criteria_indices &= criterion_indices

        return sum(float(sum_range_floats[idx]) for idx in criteria_indices)

    def evaluate_averageif(self, averageif_content: str) -> float | None:
        """Evaluate an AVERAGEIF function given its content and a spreadsheet context."""
        tokens = split_arguments(averageif_content)
        range_part = tokens[0].strip()
        criterion = tokens[1].strip().strip('"')
        average_range_part = tokens[2].strip() if len(tokens) > 2 else range_part

        range_ = self.get_range_values(range_part)
        average_range = self.get_range_values(average_range_part)

        matching_values = []
        for idx, value in enumerate(range_):
            if value is not None and self.evaluate_criterion(value, criterion):
                matching_values.append(average_range[idx])
        return sum(matching_values) / len(matching_values) if matching_values else np.nan

    def replace_functions(self, formula: str, function_to_evaluator: dict[str, Callable[[str], float]]) -> str:
        """Replace functions in the formula with their computed values.

        Assumes there are no nested functions in the formula.
        """
        start_idx = 0
        replacements = []

        while True:
            found = False
            for func_name, evaluator in function_to_evaluator.items():
                if func_name in formula[start_idx:]:
                    func_idx = formula.index(func_name, start_idx)
                    open_paren_idx = formula.index("(", func_idx)
                    close_paren_idx = formula.index(")", open_paren_idx)
                    func_content = formula[open_paren_idx + 1 : close_paren_idx]
                    value = evaluator(func_content)
                    replacements.append((func_idx, close_paren_idx + 1, str(value)))
                    start_idx = close_paren_idx + 1
                    found = True
                    break
            if not found:
                break

        for start, end, value in reversed(replacements):
            formula = formula[:start] + value + formula[end:]
        return formula

    def evaluate_formula(self, formula: str) -> list[str]:
        """Evaluate an excel-like formula given a list of tokens and a spreadsheet context.
        Use the shunting yard algorithm to convert infix to postfix notation.
        Limitations: doesn't handle nested functions.
        """
        functions = {
            "SUMIFS": self.evaluate_sumifs,
            "AVERAGEIF": self.evaluate_averageif,
        }
        formula = self.replace_functions(formula, functions)
        tokens = tokenize(formula[1:] if formula.startswith("=") else formula)
        precedence = defaultdict(lambda: 3)
        for key, value in {"+": 1, "-": 1, "*": 2, "/": 2}.items():
            precedence[key] = value
        output = []
        operators = []
        for token in tokens:
            if is_float(token):
                output.append(float(token))
            elif is_cell_or_range_reference(token):
                value = cast(float, cast(Cell, self[token]).value)
                output.append(value)
            elif token in precedence:
                while operators and operators[-1] != "(" and precedence[operators[-1]] >= precedence[token]:
                    output.append(operators.pop())
                operators.append(token)
            elif token == "(":
                operators.append(token)
            elif token == ")":
                while operators and operators[-1] != "(":
                    output.append(operators.pop())
                operators.pop()  # Remove the "("
        while operators:
            output.append(operators.pop())
        # Evaluate postfix expression
        stack = []
        for token in output:
            if isinstance(token, float):
                stack.append(token)
            else:
                b = stack.pop()
                a = stack.pop()
                if token == "+":
                    stack.append(a + b)
                elif token == "-":
                    stack.append(a - b)
                elif token == "*":
                    stack.append(a * b)
                elif token == "/":
                    stack.append(a / b if b != 0 else np.nan)
        return stack[0]

    def __getitem__(self, cell_range: str | slice) -> Cell | tuple[tuple[Cell]]:
        if isinstance(cell_range, str):
            cell = self.debug_info_sheet[cell_range]
            if cell.data_type == "f" and cell.coordinate not in self.evaluated_cells:
                evaluated_value = self.evaluate_formula(cell.value)
                self.evaluated_cells[cell.coordinate] = evaluated_value
                cell.value = evaluated_value
            return cell
        if isinstance(cell_range, slice):
            cells = self.debug_info_sheet[cell_range]
            for row in cells:
                for cell in row:
                    if cell.data_type == "f" and cell.coordinate not in self.evaluated_cells:
                        evaluated_value = self.evaluate_formula(cell.value)
                        self.evaluated_cells[cell.coordinate] = evaluated_value
                        cell.value = evaluated_value
            return cells
        raise TypeError("Invalid cell range type.")

    def get_metrics(self) -> dict[str, dict[str, float]]:
        metrics = {}
        metrics["property_precision"] = {}
        metrics["property_recall"] = {}
        formula_cells = [
            cell for row in self.debug_info_sheet.iter_rows(min_row=2) for cell in row if cell.data_type == "f"
        ]
        for precision_cell, recall_cell in zip(formula_cells[0::2], formula_cells[1::2], strict=True):
            property_id_cell = cast(Cell, precision_cell).offset(column=-1)
            property_id = property_id_cell.value
            precision = cast(float, cast(Cell, self[precision_cell.coordinate]).value)
            recall = cast(float, cast(Cell, self[recall_cell.coordinate]).value)
            if property_id != "Average":
                metrics["property_precision"][property_id] = precision if not np.isnan(precision) else None
                metrics["property_recall"][property_id] = recall if not np.isnan(recall) else None
            else:
                metrics["avg_property_precision"] = precision if not np.isnan(precision) else None
                metrics["avg_property_recall"] = recall if not np.isnan(recall) else None
        return metrics


def test_debug_output_consistency(debug_data_path: Path):
    """Test that debug output is consistent with evaluation metrics."""
    # Run evaluation to produce debug output
    run_evaluation(debug_data_path)

    # Load evaluation metrics
    metrics_file = debug_data_path / "output" / "metrics.json"
    with metrics_file.open("r", encoding="utf-8") as f:
        metrics = json.load(f)

    # Load debug output
    debug_output_file = debug_data_path / "output" / "debug_output" / "debug_info.xlsx"
    # Extract metrics from debug output
    excel_metrics = MetricsSpreadsheet(str(debug_output_file)).get_metrics()
    # Compare metrics
    assert_dicts_almost_equal(
        metrics["aesop"]["dataset_metrics"]["property_precision"], excel_metrics["property_precision"]
    )
    assert_dicts_almost_equal(metrics["aesop"]["dataset_metrics"]["property_recall"], excel_metrics["property_recall"])
    assert (
        abs(metrics["aesop"]["dataset_metrics"]["avg_property_precision"] - excel_metrics["avg_property_precision"])
        < 1e-6
    )
    assert abs(metrics["aesop"]["dataset_metrics"]["avg_property_recall"] - excel_metrics["avg_property_recall"]) < 1e-6


def run_evaluation(debug_data_path: Path):
    benchmark = mskebab.Benchmark(config_path=debug_data_path / "tasks.json", root_for_relative_paths=debug_data_path)
    task_instance = benchmark.tasks_by_name["Extraction-Debug-Info-Test"]
    predictions_file = debug_data_path / "predictions.jsonl"
    output_dir = debug_data_path / "output"
    output_dir.mkdir(parents=True, exist_ok=True)
    task_instance.evaluate(predictions_file, result_output_path=output_dir / "metrics.json")
